"""
승강시설 장애분석 데이터 정제 모듈
입력: 장애신고 xlsx + 장애분류 xlsx
출력: 처리된 엑셀(4개 시트) bytes
"""
import io
import copy
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ── 상수 ──────────────────────────────────────────────────
AC_DELETE = ['이물질','조명','데마','콤','전도','끼임','누수','정전','청소','부주의',
             '충격','화재','멀티포스트','스피커','승차감','버튼','비고장','일정수립']

EL_PRIORITY = ['안전스위치','제어장치','속도조정장치','전장품','층감지장치',
               '동력장치','동력전달장치','제동장치','도어','부속장치']
ES_PRIORITY = ['안전스위치','동력전달장치','제동장치','제어장치','속도조정장치',
               '전장품','동력장치','스텝콤','핸드레일','멀티포스트','부속장치']

# 최종 컬럼 너비
COL_WIDTHS = {
    '순번':4.625,'발생일자':10.125,'시간':6.0,'장애유형':14.125,
    '호선':7.375,'역사':8.75,'옥내/옥외':10,'신고내역':40,
    '완료일자':10.125,'완료시간':7.375,'총복구시간':8.875,'조치내역':40,'장애종류':14
}
FINAL_COLS = ['순번','발생일자','시간','장애유형','호선','역사','옥내/옥외',
              '신고내역','완료일자','완료시간','총복구시간','조치내역','장애종류']

def norm(s):
    if s is None: return ''
    return str(s).replace(' ','').replace('\u3000','').replace('\xa0','').strip()

# ── 컬럼 인덱스 동적 파악 ─────────────────────────────────
def get_col_map(ws):
    """2행 기준으로 컬럼명 → 인덱스(1-based) 매핑"""
    col_map = {}
    for c in range(1, ws.max_column+1):
        v = ws.cell(2, c).value
        if v: col_map[str(v).strip()] = c
    return col_map

# ── 필터링 ────────────────────────────────────────────────
def should_delete(row_vals, g_idx, ac_idx):
    g  = norm(row_vals[g_idx-1])
    ac = norm(row_vals[ac_idx-1])
    if '기타장애' in g or '인적장애' in g: return True
    for w in AC_DELETE:
        if norm(w) in ac: return True
    if '소음' in ac and '장력조정' in ac: return True
    return False

# ── 룩업 딕셔너리 구축 ────────────────────────────────────
def build_div(ws, col_sets):
    d = {}
    for r in range(2, ws.max_row+1):
        for (c1,c2,c3) in col_sets:
            sta = ws.cell(r,c1).value
            ho  = ws.cell(r,c2).value
            typ = ws.cell(r,c3).value
            if sta and ho and typ:
                try:    d[(str(sta).strip(), str(int(ho)))] = typ
                except: d[(str(sta).strip(), str(ho).strip())] = typ
    return d

def build_fault_dict(ws):
    headers = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
    fd = {h:[] for h in headers if h}
    for r in range(2, ws.max_row+1):
        for ci,h in enumerate(headers):
            if h:
                v = ws.cell(r, ci+1).value
                if v: fd[h].append(norm(str(v)))
    return fd

# ── 옥내/옥외 ─────────────────────────────────────────────
def get_io(mgmt_no, station, div_dict):
    parts = str(mgmt_no or '').split('-')
    try:    ho = str(int(parts[-1]))
    except: ho = parts[-1] if parts else ''
    return div_dict.get((str(station or '').strip(), ho), '')

# ── 장애종류 ──────────────────────────────────────────────
def get_fault(ac_val, fault_dict, priority):
    ac = norm(ac_val)
    matched = []
    for cat, kws in fault_dict.items():
        for kw in kws:
            if kw and kw in ac:
                matched.append(cat); break
    if not matched: return None
    matched.sort(key=lambda c: priority.index(c) if c in priority else 999)
    return matched[0]

# ── 시트 생성 ─────────────────────────────────────────────
def write_sheet(wb, name, rows):
    ws = wb.create_sheet(name)
    hdr_font  = Font(name='Arial', size=9, bold=True)
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    dat_align_c = Alignment(horizontal='center', vertical='top', wrap_text=True)
    dat_align_l = Alignment(horizontal='left',   vertical='top', wrap_text=True)
    dat_font = Font(name='Arial', size=9)

    # 헤더
    for ci, col in enumerate(FINAL_COLS, 1):
        c = ws.cell(1, ci, col)
        c.font = hdr_font; c.alignment = hdr_align
    ws.row_dimensions[1].height = 20

    # 데이터
    for ri, row in enumerate(rows, 2):
        for ci, col in enumerate(FINAL_COLS, 1):
            val = row.get(col, '')
            c = ws.cell(ri, ci, val)
            c.font = dat_font
            c.alignment = dat_align_l if col in ('신고내역','조치내역') else dat_align_c
        ws.row_dimensions[ri].height = 33.75

    # 열 너비
    for ci, col in enumerate(FINAL_COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 12)

    return ws

# ── 메인 파이프라인 ───────────────────────────────────────
def run_pipeline(fault_bytes, cls_bytes, progress_cb=None):
    def prog(msg, pct):
        if progress_cb: progress_cb(msg, pct)

    prog('파일 읽는 중...', 5)
    wb_src = load_workbook(io.BytesIO(fault_bytes))
    wb_cls = load_workbook(io.BytesIO(cls_bytes))

    # 시트명 유연 처리 (sheet1 또는 장애신고)
    src_sheet = None
    for sn in wb_src.sheetnames:
        if '장애신고' in sn or sn.lower() == 'sheet1':
            src_sheet = wb_src[sn]; break
    if src_sheet is None:
        src_sheet = wb_src[wb_src.sheetnames[0]]

    prog('컬럼 구조 파악 중...', 10)
    col_map = get_col_map(src_sheet)

    # 컬럼 인덱스
    G  = col_map.get('장애유형', 7)
    H  = col_map.get('호선', 8)
    I  = col_map.get('역사', 9)
    J  = col_map.get('관리번호', 10)
    P  = col_map.get('신고내역', 16)
    W  = col_map.get('완료일자', 23)
    X  = col_map.get('완료시간', 24)
    AB = col_map.get('총복구시간', 28)
    AC = col_map.get('조치내역', 29)
    C  = col_map.get('발생일자', 3)
    D  = col_map.get('시간', 4)

    prog('분류 데이터 로드 중...', 15)
    el_div = build_div(wb_cls['EL구분'], [(1,2,3),(5,6,7)])
    es_div = build_div(wb_cls['ES구분'], [(1,2,3),(5,6,7)])
    el_fd  = build_fault_dict(wb_cls['EL장애'])
    es_fd  = build_fault_dict(wb_cls['ES장애'])

    prog('데이터 필터링 중...', 25)
    all_rows = []
    for r in range(3, src_sheet.max_row+1):
        vals = [src_sheet.cell(r,c).value for c in range(1, src_sheet.max_column+1)]
        if any(v is not None for v in vals):
            if not should_delete(vals, G, AC):
                all_rows.append(vals)

    prog('EL/ES/호선 분리 중...', 40)
    el1,el2,es1,es2 = [],[],[],[]
    for vals in all_rows:
        mn   = norm(vals[J-1])
        line = str(vals[H-1] or '')
        is1  = '1호선' in line
        is2  = '2호선' in line
        if 'EL' in mn:
            if is1: el1.append(vals)
            elif is2: el2.append(vals)
        elif 'ES' in mn:
            if is1: es1.append(vals)
            elif is2: es2.append(vals)

    prog('장애종류 분류 중...', 60)

    def make_rows(data_list, div_dict, fd, priority):
        result = []
        seq = 1
        for vals in data_list:
            ac_val = vals[AC-1]
            fault  = get_fault(ac_val, fd, priority)
            if fault is None: continue
            io_val = get_io(vals[J-1], vals[I-1], div_dict)
            result.append({
                '순번':       seq,
                '발생일자':   vals[C-1],
                '시간':       vals[D-1],
                '장애유형':   vals[G-1],
                '호선':       vals[H-1],
                '역사':       vals[I-1],
                '옥내/옥외':  io_val,
                '신고내역':   vals[P-1],
                '완료일자':   vals[W-1],
                '완료시간':   vals[X-1],
                '총복구시간': vals[AB-1],
                '조치내역':   ac_val,
                '장애종류':   fault,
            })
            seq += 1
        return result

    r_el1 = make_rows(el1, el_div, el_fd, EL_PRIORITY)
    r_el2 = make_rows(el2, el_div, el_fd, EL_PRIORITY)
    r_es1 = make_rows(es1, es_div, es_fd, ES_PRIORITY)
    r_es2 = make_rows(es2, es_div, es_fd, ES_PRIORITY)

    prog('엑셀 시트 생성 중...', 80)
    wb_out = Workbook()
    wb_out.remove(wb_out.active)
    write_sheet(wb_out, '1호선EL', r_el1)
    write_sheet(wb_out, '2호선EL', r_el2)
    write_sheet(wb_out, '1호선ES', r_es1)
    write_sheet(wb_out, '2호선ES', r_es2)

    out = io.BytesIO()
    wb_out.save(out)
    out.seek(0)

    stats = {
        '1호선EL': len(r_el1),
        '2호선EL': len(r_el2),
        '1호선ES': len(r_es1),
        '2호선ES': len(r_es2),
    }

    prog('완료', 100)
    return out.read(), stats
